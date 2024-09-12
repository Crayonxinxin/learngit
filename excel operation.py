'''
Description: 
Author: CrayonXiaoxin
Date: 2024-09-10 09:49:17
LastEditTime: 2024-09-12 15:04:06
LastEditors:  
'''
import os
import openpyxl

if __name__ == '__main__':
    path = os.path.dirname(__file__)
    os.chdir(path)
    print(path)
    # 检查文件是否存在
    file_name = input("请输入文件名：")
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"文件 {file_name} 不存在！")

    workbook = openpyxl.load_workbook(file_name)
    print(workbook.sheetnames)
    sheet1 = input("请输入主表名字：")
    # 获取 Sheet1（假设其名字是 '1'）
    if sheet1 not in workbook.sheetnames:
        raise ValueError("找不到名为 %s 的工作表"%sheet1)

    sheet = workbook[sheet1]
    rows = sheet.max_row
    column = sheet.max_column

    # 获取表头
    topName = []
    for i in range(1, column+1):
        topName.append(sheet.cell(1, i).value)
    names = []
    # 自定义分类名称
    for i in range(1,1000):
        name = input("请输入要整理的名字：(如果想结束请输入#)")
        if name == '#':
            break
        names.append(name)
    

    # 创建新的 Sheet
    for name in names:
        if name not in workbook.sheetnames:
            workbook.create_sheet(name)

    # 按分类写入数据
    for i in range(2, rows+1):  # 从第2行开始读取，因为第1行是表头
        name = sheet.cell(i, 2).value
        if name in names:
            viceSheet = workbook[name]
            values = [sheet.cell(i, j).value for j in range(1, column+1)]
            vicerows = viceSheet.max_row
            if vicerows == 0:
                # 如果新建的表是空的，首先写入表头
                for l in range(1, column+1):
                    viceSheet.cell(1, l, topName[l-1])
                vicerows = 1
            for k in range(1, column+1):
                viceSheet.cell(vicerows + 1, k, values[k-1])

    # 保存并关闭文件
    workbook.save(file_name)
    workbook.close()
